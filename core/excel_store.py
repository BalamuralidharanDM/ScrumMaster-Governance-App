from __future__ import annotations
from contextlib import contextmanager
from datetime import datetime
from pathlib import Path
from typing import Dict, List
import re
import pandas as pd
from filelock import FileLock
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]; DATA_FILE=ROOT/'data'/'scrum_master_backend.xlsx'; BACKUP_DIR=ROOT/'backups'; LOCK_FILE=ROOT/'data'/'scrum_master_backend.lock'
SHEET_COLUMNS={
'App Tasks':['Task ID','Project','Epic','Sprint','Task','Owner','Peer QA','Status','Priority','Start Date','End Date','Duration Days','Progress %','Expected %','Variance %','Risk','Dependency','Milestone','Last Updated','Comments'],
'App RAID':['RAID ID','Type','Description','Impact','Probability','Owner','Mitigation','Due Date','Status','Escalation Trigger','Last Updated'],
'App Resources':['Name','Level','Location','Skill Set','Role','Capacity %','Available From','Available To'],
'App Projects':['Project ID','Project Name','Description','Project Manager','Status','Priority','Start Date','End Date','Active'],
'App Epics':['Epic ID','Project Name','Epic Name','Description','Owner','Status','Priority','Active'],
'App Sprints':['Sprint ID','Project Name','Sprint','Start Date','End Date','Goal','Status','Active'],
'App Owners':['Owner ID','Owner Name','Email','Role','Location','Active'],
'App Statuses':['Status ID','Status Name','Category','Sort Order','Active'],
'App Priorities':['Priority ID','Priority Name','Sort Order','Active'],
'App Decisions':['Decision ID','Date','Topic','Decision','Owner','Status','Impact']}
class ExcelStore:
    def __init__(self,path:Path=DATA_FILE): self.path=Path(path); self.path.parent.mkdir(parents=True,exist_ok=True); BACKUP_DIR.mkdir(parents=True,exist_ok=True)
    @contextmanager
    def locked(self):
        with FileLock(str(LOCK_FILE),timeout=30): yield
    def backup(self):
        stamp=datetime.now().strftime('%Y%m%d_%H%M%S_%f'); target=BACKUP_DIR/f'scrum_master_backend_{stamp}.xlsx'
        if self.path.exists(): target.write_bytes(self.path.read_bytes())
        return target
    def read_sheet(self,sheet_name:str,header:int=0):
        with self.locked():
            if not self.path.exists(): return pd.DataFrame(columns=SHEET_COLUMNS.get(sheet_name,[]))
            try:return pd.read_excel(self.path,sheet_name=sheet_name,header=header,engine='openpyxl')
            except ValueError:return pd.DataFrame(columns=SHEET_COLUMNS.get(sheet_name,[]))
    def list_sheets(self)->List[str]:
        with self.locked():
            if not self.path.exists():return []
            return load_workbook(self.path,read_only=True).sheetnames
    def write_sheet(self,sheet_name:str,df:pd.DataFrame):
        with self.locked():
            self.backup(); mode='a' if self.path.exists() else 'w'; kwargs=dict(engine='openpyxl',mode=mode)
            if mode=='a':kwargs['if_sheet_exists']='replace'
            with pd.ExcelWriter(self.path,**kwargs) as writer:df.to_excel(writer,sheet_name=sheet_name,index=False)
    def ensure_app_sheets(self):
        existing=set(self.list_sheets())
        for n,c in SHEET_COLUMNS.items():
            if n not in existing:self.write_sheet(n,pd.DataFrame(columns=c))
    @staticmethod
    def next_id(df,column,prefix,width=4):
        n=0
        if column in df:
            for v in df[column].dropna().astype(str):
                m=re.search(r'(\d+)$',v.strip())
                if m:n=max(n,int(m.group(1)))
        return f'{prefix}-{n+1:0{width}d}'
def to_numeric(series,default=0):return pd.to_numeric(series,errors='coerce').fillna(default)
